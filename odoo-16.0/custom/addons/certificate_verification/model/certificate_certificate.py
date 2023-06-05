import base64
import datetime
from io import BytesIO
from odoo import api, fields, models
from openpyxl import Workbook
from openpyxl.styles import Font, DEFAULT_FONT


class Certificate(models.Model):
    _name = "certificate.certificate"
    _description = "Certificate"

    name = fields.Char("Certificate No.", readonly=True, copy=False)
    partner_id = fields.Many2one('res.partner', string="Student")
    sale_order_id = fields.Many2one('sale.order', string="Registration No", readonly=True)
    course_id = fields.Many2one('product.template', related="sale_order_id.order_line.product_template_id",
                                string="Course Name")
    street = fields.Char('Street', related="partner_id.street")
    street2 = fields.Char('Street2', related="partner_id.street2")
    zip = fields.Char('Zip', related="partner_id.zip")
    city = fields.Char('City', related="partner_id.city")
    state_id = fields.Many2one("res.country.state", string='State', related="partner_id.state_id")
    country_id = fields.Many2one('res.country', string='Country', related="partner_id.country_id")
    address = fields.Char(compute="_compute_address", string="Address", )
    phone = fields.Char(string="Phone", related="partner_id.phone")
    course_status = fields.Selection([
        ('new', 'New'),
        ('not_started', 'Not Started'),
        ('completed', 'Completed'),
        ('running', 'Running')], compute="compute_course_status", string="Course Status")
    fee_status = fields.Selection([
        ('upselling', "Upselling Opportunity"),
        ('invoiced', "Fully Invoiced"),
        ('to invoice', "To Invoice"),
        ('no', "Nothing to Invoice"),
    ], related='sale_order_id.invoice_status')

    course_start_datetime = fields.Datetime(string='Course Starting')
    course_end_datetime = fields.Datetime(string='Course Ending')

    _sql_constraints = [
        (
            'check_course_date',
            'CHECK(course_start_datetime < course_end_datetime)',
            "Please select the correct date.",
        ),
    ]

    def _compute_address(self):
        for rec in self:
            if rec.partner_id:
                street = rec.street + ', ' if rec.street else ''
                street2 = rec.street2 + ', ' if rec.street2 else ''
                city = rec.city + ', ' if rec.city else ''
                zip = rec.zip + ', ' if rec.zip else ''
                state = rec.state_id.name + ', ' if rec.state_id else ''
                country = rec.country_id.name + ', ' if rec.country_id else ''
                rec.address = street + street2 + city + zip + state + country
            else:
                rec.address = ''

    def compute_course_status(self):
        now = datetime.datetime.now()
        for rec in self:
            if rec.course_start_datetime and rec.course_end_datetime:
                if now < rec.course_start_datetime:
                    rec.course_status = 'not_started'
                elif now > rec.course_end_datetime:
                    rec.course_status = 'completed'
                else:
                    rec.course_status = 'running'
            else:
                rec.course_status = 'new'

    @api.model
    def create(self, vals):
        vals['name'] = self.env['ir.sequence'].next_by_code('certificate.sequence') or 'New'
        res = super(Certificate, self).create(vals)
        return res

    def generate_excel(self):
        DEFAULT_FONT.name = "Arial"
        DEFAULT_FONT.size = 10
        workbook = Workbook(write_only=False)

        heading_font = Font(size=10, bold=True, name='Arial')
        column_font = Font(size=10, name='Arial')
        worksheet = workbook.active

        worksheet.cell(row=1, column=1, value="Certificate No.").font = heading_font
        worksheet.cell(row=1, column=2, value="Student Name").font = heading_font
        worksheet.cell(row=1, column=3, value="Address").font = heading_font
        worksheet.cell(row=1, column=4, value="Phone").font = heading_font
        worksheet.cell(row=1, column=5, value="Registration No.").font = heading_font
        worksheet.cell(row=1, column=6, value="Course Name").font = heading_font
        worksheet.cell(row=1, column=7, value="Course Status").font = heading_font
        worksheet.cell(row=1, column=8, value="Fees Status").font = heading_font

        if self:
            all_data = self

        else:
            all_data = self.env['certificate.certificate'].search([])
        i = 2
        for data in all_data:
            worksheet.cell(row=i, column=1, value=data.name).font = column_font
            worksheet.cell(row=i, column=2, value=data.partner_id.name).font = column_font
            worksheet.cell(row=i, column=3, value=data.address).font = column_font
            worksheet.cell(row=i, column=4, value=data.phone).font = column_font
            worksheet.cell(row=i, column=5, value=data.sale_order_id.name).font = column_font
            worksheet.cell(row=i, column=6, value=data.course_id.name).font = column_font
            worksheet.cell(row=i, column=7, value=data.course_status).font = column_font
            worksheet.cell(row=i, column=8, value=data.fee_status).font = column_font

            i += 1
        fp = BytesIO()
        workbook.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        fp.close()
        ss = self.env['ir.attachment'].create({'name': 'certificate', 'datas': excel_file})

        return {
            'view_mode': 'form',
            'res_model': 'ir.attachment',
            'res_id': ss.id,
            'views': [(self.env.ref('certificate_verification.excel_download').id, 'form')],
            'type': 'ir.actions.act_window',
            'context': self.env.context,
            'target': 'new',
        }
